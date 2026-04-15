"""
survey联合判定脚本
对Excel信息表中的所有样本，依次执行kmer峰型判断和NT比对污染判断，输出综合结果。
"""
import os
import glob
import pandas as pd
from kmer_judge import main_dual
from nt_judge import judge_nt_contamination


EXCEL_PATH = '/data/work/zhurui/survey_rec/data/survey 信息表.xlsx'
SHEET_NAME = '处理后信息'
BASE_PATH = '/data/work/zhurui/survey_rec/data/shenshaoqi_data/survey1'

KMER_PATTERN_CN = {
    'diploid_homo': '纯合二倍体',
    'diploid_hetero': '杂合二倍体',
    'triploid': '三倍体',
    'high_hetero_diplo': '高杂合二倍体',
    'high_repetitive_diplo': '高重复二倍体',
    'tetraploid': '四倍体',
    'no_peak': '无峰',
    'unknown': '未知',
    'error': '错误',
    'missing': '缺失',
}


def get_real_path(raw_path):
    """将原始路径转换为本地真实路径"""
    parts = str(raw_path).strip().split('/')
    for i, p in enumerate(parts):
        if p.startswith('X101SC'):
            return os.path.join(BASE_PATH, *parts[i:])
    return None


def get_sample_name(sample_dir):
    """从样本目录名提取样本名（最后一个_后的部分）"""
    dirname = os.path.basename(sample_dir)
    if '_' in dirname:
        return dirname.split('_', 1)[1]
    return dirname


def find_kmer_files(sample_dir, sample_name):
    """查找kmer文件，支持模糊匹配"""
    spe = os.path.join(sample_dir, f'{sample_name}.17merFreq.SpeFreq.cut')
    num = os.path.join(sample_dir, f'{sample_name}.17merFreq.NumFreq.cut')
    if os.path.exists(spe) and os.path.exists(num):
        return spe, num
    # 模糊匹配
    spe_list = glob.glob(os.path.join(sample_dir, '*.SpeFreq.cut'))
    num_list = glob.glob(os.path.join(sample_dir, '*.NumFreq.cut'))
    if spe_list and num_list:
        return spe_list[0], num_list[0]
    return None, None


def judge_sample(sample_dir, verbose=False):
    """
    对单个样本执行联合判定
    返回 dict: {
        'sample_dir', 'sample_name', 'target_species',
        'kmer_pattern', 'kmer_normal', 'kmer_detail',
        'nt_score', 'nt_level', 'ntcls_score', 'ntspe_score',
        'ntcls_detail', 'ntspe_detail',
        'final_level', 'error'
    }
    """
    result = {'sample_dir': sample_dir, 'error': None}
    sample_name = get_sample_name(sample_dir)
    result['sample_name'] = sample_name

    # 读取物种名（来自ntcls文件第一列）
    ntcls_path = os.path.join(sample_dir, 'all.ntcls.xls')
    ntspe_path = os.path.join(sample_dir, 'all.ntspe.xls')

    try:
        df_cls = pd.read_csv(ntcls_path, sep='\t')
        target_species = str(df_cls.iloc[0]['Sample name'])
    except Exception as e:
        result['error'] = f'读取物种名失败: {e}'
        result['final_level'] = 'fail'
        return result

    result['target_species'] = target_species

    # 收集判定依据
    basis_parts = [f"物种: {target_species}"]

    # kmer判断
    spe_file, num_file = find_kmer_files(sample_dir, sample_name)
    if spe_file and num_file:
        try:
            kmer_res = main_dual(spe_file, num_file, verbose=verbose)
            result['kmer_pattern'] = kmer_res['pattern']
            result['kmer_normal'] = kmer_res['is_normal']
            result['kmer_detail'] = kmer_res['detail']
            basis_parts.append(f"kmer: {kmer_res['detail']}")
        except Exception as e:
            result['kmer_pattern'] = 'error'
            result['kmer_normal'] = False
            result['kmer_detail'] = f'kmer判断异常: {e}'
            basis_parts.append(f"kmer: 判断异常-{e}")
    else:
        result['kmer_pattern'] = 'missing'
        result['kmer_normal'] = False
        result['kmer_detail'] = 'kmer文件不存在'
        basis_parts.append("kmer: 文件不存在")

    # NT比对判断
    try:
        nt_res = judge_nt_contamination(ntcls_path, ntspe_path, target_species)
        result['nt_score'] = nt_res['nt_score']
        result['nt_level'] = nt_res['nt_level']
        result['ntcls_score'] = nt_res['ntcls_score']
        result['ntspe_score'] = nt_res['ntspe_score']
        result['ntcls_detail'] = nt_res.get('ntcls_detail', '')
        result['ntspe_detail'] = nt_res.get('ntspe_detail', '')
        result['ntcls_top1_pass'] = nt_res.get('ntcls_top1_pass', False)
        result['ntcls_contamination_pass'] = nt_res.get('ntcls_contamination_pass', False)
        result['ntspe_contamination_pass'] = nt_res.get('ntspe_contamination_pass', False)
        basis_parts.append(f"NT大类: {nt_res.get('ntcls_detail', '')}")
        basis_parts.append(f"NT小类: {nt_res.get('ntspe_detail', '')}")
        basis_parts.append(f"NT总分={nt_res['nt_score']}, 等级={nt_res['nt_level']}")
    except Exception as e:
        result['nt_score'] = 0
        result['nt_level'] = 'fail'
        result['ntcls_score'] = 0
        result['ntspe_score'] = 0
        result['ntcls_detail'] = f'NT判断异常: {e}'
        result['ntspe_detail'] = ''
        result['ntcls_top1_pass'] = False
        result['ntcls_contamination_pass'] = False
        result['ntspe_contamination_pass'] = False
        basis_parts.append(f"NT: 判断异常-{e}")

    # 综合判定：kmer + nt_level 联合
    kmer_normal = result.get('kmer_normal', False)
    nt_level = result.get('nt_level', 'fail')
    nt_score = result.get('nt_score', 0)
    if kmer_normal:
        # kmer正常
        if nt_level in ('正常', '轻度污染'):
            result['final_level'] = '正常'
            result['should_transfer'] = '是'
            result['remark'] = ''
        elif nt_level == '重度污染':
            # 重度污染中，kmer正常但NT<=2分，不流转
            if nt_score <= 2:
                result['final_level'] = '轻度污染'
                result['should_transfer'] = '否'
                result['remark'] = 'NT得分<=2分，不建议流转'
            else:
                result['final_level'] = '轻度污染'
                result['should_transfer'] = '是'
                result['remark'] = ''
        else:  # fail
            if nt_score >= 3:
                result['final_level'] = '轻度污染'
                result['should_transfer'] = '是'
                result['remark'] = ''
            else:
                result['final_level'] = '重度污染'
                result['should_transfer'] = '否'
                result['remark'] = 'NT得分<=2分，不建议流转'
    else:
        # kmer异常
        if nt_level == '正常':
            # NT正常10分但kmer异常，不流转
            result['final_level'] = '重度污染'
            result['should_transfer'] = '否'
            result['remark'] = 'NT正常但kmer异常，不建议流转'
        else:
            result['final_level'] = 'fail'
            result['should_transfer'] = '否'
            result['remark'] = ''

    result['basis'] = '\n'.join(basis_parts)

    return result


def run_all(max_samples=None, verbose=True, output_path=None):
    """
    批量处理Excel中所有样本，结果写入新的Excel文件
    """
    from openpyxl import load_workbook

    df = pd.read_excel(EXCEL_PATH, sheet_name=SHEET_NAME)
    total = len(df) if max_samples is None else min(max_samples, len(df))

    print(f'共 {total} 个样本待处理')
    print('=' * 60)

    new_cols = ['kmer峰型', 'kmer是否正常', 'NT大类Top1判断', 'NT大类污染判断', 'NT小类污染判断', 'NT得分', 'NT等级', '综合判定', '是否流转', '判定依据', '备注']
    for col in new_cols:
        df[col] = ''

    for i in range(total):
        raw_path = df.iloc[i]['路径']
        sample_dir = get_real_path(raw_path)

        print(f'\n[{i+1}/{total}] {os.path.basename(sample_dir) if sample_dir else raw_path}')
        print(f'  路径: {sample_dir if sample_dir else raw_path}')

        if not sample_dir or not os.path.isdir(sample_dir):
            print('  路径不存在，跳过')
            df.at[i, '综合判定'] = '路径不存在'
            continue

        res = judge_sample(sample_dir, verbose=verbose)

        if verbose:
            print(f"  判定结果: {res}")

        df.at[i, 'kmer峰型'] = KMER_PATTERN_CN.get(res.get('kmer_pattern', ''), res.get('kmer_pattern', ''))
        df.at[i, 'kmer是否正常'] = '正常' if res.get('kmer_normal') else '异常'
        df.at[i, 'NT大类Top1判断'] = 'pass' if res.get('ntcls_top1_pass') else 'fail'
        df.at[i, 'NT大类污染判断'] = 'pass' if res.get('ntcls_contamination_pass') else 'fail'
        df.at[i, 'NT小类污染判断'] = 'pass' if res.get('ntspe_contamination_pass') else 'fail'
        df.at[i, 'NT得分'] = res.get('nt_score', 0)
        df.at[i, 'NT等级'] = res.get('nt_level', '')
        df.at[i, '综合判定'] = res.get('final_level', 'fail')
        df.at[i, '是否流转'] = res.get('should_transfer', '否')
        df.at[i, '判定依据'] = res.get('basis', '')
        df.at[i, '备注'] = res.get('remark', '')

        kmer_str = f"{res.get('kmer_pattern','?')}({'正常' if res.get('kmer_normal') else '异常'})"
        nt_str = f"得分={res.get('nt_score', 0)} 等级={res.get('nt_level', '?')}"
        print(f'  物种: {res.get("target_species","?")}')
        print(f'  kmer={kmer_str} | NT: {nt_str} | 综合: {df.at[i, "综合判定"]}')

    # 写入新Excel文件
    if output_path is None:
        base, ext = os.path.splitext(EXCEL_PATH)
        output_path = base + '_结果' + ext

    wb = load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]

    # 找到已有列数，追加新列头
    header = [cell.value for cell in ws[1]]
    for col in new_cols:
        if col not in header:
            header.append(col)
            ws.cell(row=1, column=len(header), value=col)

    # 写入数据
    for i in range(total):
        for col in new_cols:
            col_idx = header.index(col) + 1
            ws.cell(row=i + 2, column=col_idx, value=df.at[i, col])

    wb.save(output_path)

    normal_count = sum(1 for i in range(total) if df.at[i, '综合判定'] == '正常')
    print('\n' + '=' * 60)
    print(f'处理完成: {total} 个样本，正常 {normal_count} 个，异常 {total - normal_count} 个')
    print(f'结果已写入: {output_path}')

    return df


if __name__ == '__main__':
    import argparse

    parser = argparse.ArgumentParser(description='Survey联合判定')
    parser.add_argument('--max', type=int, default=None, help='最多处理样本数（测试用）')
    parser.add_argument('--verbose', action='store_true', default=True, help='打印详细过程')
    args = parser.parse_args()

    run_all(max_samples=args.max, verbose=args.verbose)
