"""
仅进行kmer峰型判断的脚本
读取Excel信息表中的所有样本，执行kmer峰型判断，输出结果到Excel
"""
import os
import glob
import pandas as pd
from kmer_judge import main_dual


EXCEL_PATH = '/data/work/zhurui/survey_rec/data/survey 信息表.xlsx'
SHEET_NAME = '处理后信息'
BASE_PATH = '/data/work/zhurui/survey_rec/data/shenshaoqi_data/survey1'

KMER_PATTERN_CN = {
    'diploid_homo': '纯合二倍体',
    'diploid_hetero': '杂合二倍体',
    'triploid': '三倍体',
    'high_hetero_diplo': '高杂合二倍体',
    'high_repetitive_diplo': '高重复二倍体',
    'tetraploid': '四倍体',
    'suspected_polyploid': '疑似多倍体',
    'no_peak': '无峰',
    'unknown': '未知',
    'error': '错误',
    'missing': '缺失',
}


def to_arabic_ploidy(ploidy_text):
    """将中文/英文倍型描述归一为阿拉伯数字倍型（如 2倍体），无法识别返回空字符串"""
    s = str(ploidy_text).strip()
    if not s:
        return ''

    mapping = {
        '二倍体': '2倍体',
        '纯合二倍体': '2倍体',
        '杂合二倍体': '2倍体',
        '高杂合二倍体': '2倍体',
        '高重复二倍体': '2倍体',
        'diploid': '2倍体',
        'diploid_homo': '2倍体',
        'diploid_hetero': '2倍体',
        'high_hetero_diplo': '2倍体',
        'high_repetitive_diplo': '2倍体', 
        '三倍体': '3倍体',
        'triploid': '3倍体',
        '四倍体': '4倍体',
        'tetraploid': '4倍体',
    }
    return mapping.get(s, '')


def get_real_path(raw_path):
    """将原始路径转换为本地真实路径"""
    parts = str(raw_path).strip().split('/')
    for i, p in enumerate(parts):
        if p.startswith('X101SC'):
            return os.path.join(BASE_PATH, *parts[i:])
    return None


def get_sample_name(sample_dir):
    """从样本目录名提取样本名"""
    dirname = os.path.basename(sample_dir)
    if '_' in dirname:
        return dirname.split('_', 1)[1]
    return dirname


def find_kmer_files(sample_dir, sample_name):
    """查找kmer文件"""
    spe = os.path.join(sample_dir, f'{sample_name}.17merFreq.SpeFreq.cut')
    num = os.path.join(sample_dir, f'{sample_name}.17merFreq.NumFreq.cut')
    if os.path.exists(spe) and os.path.exists(num):
        return spe, num
    spe_list = glob.glob(os.path.join(sample_dir, '*.SpeFreq.cut'))
    num_list = glob.glob(os.path.join(sample_dir, '*.NumFreq.cut'))
    if spe_list and num_list:
        return spe_list[0], num_list[0]
    return None, None


def judge_kmer_only(sample_dir, verbose=False):
    """对单个样本执行kmer判断"""
    result = {'sample_dir': sample_dir}
    sample_name = get_sample_name(sample_dir)
    result['sample_name'] = sample_name
    result['kmer_warning'] = ''

    spe_file, num_file = find_kmer_files(sample_dir, sample_name)
    if spe_file and num_file:
        try:
            kmer_res = main_dual(spe_file, num_file, verbose=verbose)
            result['kmer_pattern'] = kmer_res['pattern']
            result['kmer_normal'] = kmer_res['is_normal']
            result['kmer_detail'] = kmer_res['detail']
            result['kmer_warning'] = '；'.join(kmer_res.get('warnings', []))
        except Exception as e:
            result['kmer_pattern'] = 'error'
            result['kmer_normal'] = False
            result['kmer_detail'] = f'kmer判断异常: {e}'
            result['kmer_warning'] = ''
    else:
        result['kmer_pattern'] = 'missing'
        result['kmer_normal'] = False
        result['kmer_detail'] = 'kmer文件不存在'
        result['kmer_warning'] = ''

    return result


def run_all(max_samples=None, verbose=False, output_path=None):
    """批量处理Excel中所有样本，结果写入新的Excel文件"""
    from openpyxl import load_workbook

    df = pd.read_excel(EXCEL_PATH, sheet_name=SHEET_NAME)
    total = len(df) if max_samples is None else min(max_samples, len(df))

    print(f'共 {total} 个样本待处理')
    print('=' * 60)

    new_cols = ['kmer峰型', 'kmer是否正常', 'kmer详情', 'kmer警告信息', '是否一致', '倍型判断是否一致']
    for col in new_cols:
        df[col] = ''

    for i in range(total):
        raw_path = df.iloc[i]['路径']
        sample_dir = get_real_path(raw_path)

        print(f'\n[{i+1}/{total}] {os.path.basename(sample_dir) if sample_dir else raw_path}')

        if not sample_dir or not os.path.isdir(sample_dir):
            print('  路径不存在，跳过')
            df.at[i, 'kmer峰型'] = '路径不存在'
            continue

        res = judge_kmer_only(sample_dir, verbose=verbose)

        df.at[i, 'kmer峰型'] = KMER_PATTERN_CN.get(res.get('kmer_pattern', ''), res.get('kmer_pattern', ''))
        df.at[i, 'kmer是否正常'] = '正常' if res.get('kmer_normal') else '异常'
        df.at[i, 'kmer详情'] = res.get('kmer_detail', '')
        df.at[i, 'kmer警告信息'] = res.get('kmer_warning', '')

        original_poisson = str(df.iloc[i].get('是否符合泊松分布', '')).strip()
        kmer_result = df.at[i, 'kmer是否正常']
        is_match = (original_poisson == '是' and kmer_result == '正常') or (original_poisson == '否' and kmer_result == '异常')
        df.at[i, '是否一致'] = '是' if is_match else '否'

        if kmer_result == '正常':
            original_ploidy = str(df.iloc[i].get('物种倍性', '')).strip()
            kmer_ploidy = to_arabic_ploidy(df.at[i, 'kmer峰型'])
            if original_ploidy and kmer_ploidy:
                df.at[i, '倍型判断是否一致'] = '是' if original_ploidy == kmer_ploidy else '否'
            else:
                df.at[i, '倍型判断是否一致'] = ''
        else:
            df.at[i, '倍型判断是否一致'] = ''

        print(f'  kmer峰型: {df.at[i, "kmer峰型"]} | 是否正常: {df.at[i, "kmer是否正常"]}')
        if df.at[i, 'kmer警告信息']:
            print(f'  警告信息: {df.at[i, "kmer警告信息"]}')

    if output_path is None:
        base, ext = os.path.splitext(EXCEL_PATH)
        output_path = base + '_kmer结果' + ext

    wb = load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]

    header = [cell.value for cell in ws[1]]
    for col in new_cols:
        if col not in header:
            header.append(col)
            ws.cell(row=1, column=len(header), value=col)

    for i in range(total):
        for col in new_cols:
            col_idx = header.index(col) + 1
            ws.cell(row=i + 2, column=col_idx, value=df.at[i, col])

    wb.save(output_path)

    normal_count = sum(1 for i in range(total) if df.at[i, 'kmer是否正常'] == '正常')
    match_count = sum(1 for i in range(total) if df.at[i, '是否一致'] == '是')
    ploidy_match_count = sum(
        1 for i in range(total)
        if df.at[i, 'kmer是否正常'] == '正常' and df.at[i, '倍型判断是否一致'] == '是'
    )
    ploidy_mismatch_count = sum(
        1 for i in range(total)
        if df.at[i, 'kmer是否正常'] == '正常' and df.at[i, '倍型判断是否一致'] == '否'
    )
    print('\n' + '=' * 60)
    print(f'处理完成: {total} 个样本，正常 {normal_count} 个，异常 {total - normal_count} 个')
    print(f'一致 {match_count} 个，不一致 {total - match_count} 个')
    print(
        f'倍型一致（仅正常样本） {ploidy_match_count} 个，'
        f'倍型不一致（仅正常样本） {ploidy_mismatch_count} 个'
    )
    print(f'结果已写入: {output_path}')

    return df

"""
使用方法
conda activate zhurui_agent
python kmer_only_judge.py --max 10  # 测试前10个样本
python kmer_only_judge.py --verbose  # 打印详细过程
"""



if __name__ == '__main__':
    import argparse

    parser = argparse.ArgumentParser(description='Survey kmer判定')
    parser.add_argument('--max', type=int, default=None, help='最多处理样本数（测试用）')
    parser.add_argument('--verbose', action='store_true', help='打印详细过程')
    args = parser.parse_args()

    run_all(max_samples=args.max, verbose=args.verbose)
